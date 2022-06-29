import sqlite3
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkcalendar import *
import openpyxl
from openpyxl import Workbook, load_workbook
import os

win = tk.Tk()  # Application Name
win.title("Clinker")  # Label

DoS = ""
DoA = ""
filterDateTo = ""
filterDateFrom = ""

wb = load_workbook('Experiment.xlsx')
ws = wb.active
# print(ws['F8'].value)

# connection to database
clin = sqlite3.connect('clinker.db')

# create a cursor
c = clin.cursor()

# Excel related stuff
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet_title = 'Clinker'
# print("My sheet title: " + my_sheet_title)

def createTable():
    # create a table
    c.execute("""CREATE TABLE clinkerData (
        Date_of_sample text,
        Date_of_analysis text,
        Time text,
        Material_name text,
        Moisture text,
        LOI text,
        SiO2 text,
        Al2O3 text,
        Fe2O3 text,
        CaO text,
        MgO text,
        SO3 text,
        Na2O text,
        K2O text,
        P2O5 text,
        Mn2O3 text,
        Cl text,
        FreeCaO text,
        IR text,
        Literweight text


    )

    """)


def loadDatatoDB():

    # insert data into database table

    c.execute("INSERT INTO clinkerData VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
              (DoS.get_date(),DoA.get_date(),Time.get(),MaterialName.get(),Moisture.get(),LOI.get(),SiO2.get(),
               Al2O3.get(),Fe2O3.get(),CaO.get(),MgO.get(),SO3.get(),Na2O.get(),K2O.get(),
               P2O5.get(),Mn2O3.get(),Cl.get(),FreeCaO.get(),IR.get(),Literweight.get()))


def testingw():
    print(Moisture.get())
    mtn = MaterialName.get()
    print(mtn)


def listofcommands():
    # createTable()
    loadDatatoDB()
    testingw()

materialOptions = [
        "Clinker ypw",
        "OPC 53-Grinding",
        "OPC 53-Packing",
        "PPC-Grinding",
        "PPC-Packing",
        "Ball mill OPC",
        "Ball mill PPC",
        "Gypsum-Chemical",
        "Gypsum-Mineral",
        "Gypsum-Saltpan",
        "Gypsum-Lowgrade",
        "Clinker-MCW",
        "Clinker-Local",
        "Wet Flyash",
        "Dry Flyash",
        "Limestone",
        "Additive-1",
        "Additive-2",
        "Additive-3",
        "Sample-1",
        "Sample-2",
        "Sample-3"
    ]

def timeSelection():
    timeOptions = [
        "0:00",
        "1:00",
        "2:00",
        "3:00",
        "4:00",
        "5:00",
        "6:00",
        "7:00",
        "8:00",
        "9:00",
        "10:00",
        "11:00",
        "12:00",
        "13:00",
        "14:00",
        "15:00",
        "16:00",
        "17:00",
        "18:00",
        "19:00",
        "20:00",
        "21:00",
        "22:00",
        "23:00"
    ]


    global DoS,DoA
    DoS = Calendar(win, selectmode ='day', date_pattern="d-m-y")
    DoS.grid(column=1, row=0)
    ttk.Label(win, text="Date of Sample :").grid(column=0, row=0)
    ttk.Label(win, text="Material Name").grid(column=0, row=1)
    ttk.Label(win, text="Moisture").grid(column=0, row=2)
    ttk.Label(win, text="LOI").grid(column=0, row=3)
    ttk.Label(win, text="SiO₂").grid(column=0, row=4)
    ttk.Label(win, text="Al₂O₃").grid(column=0, row=5)
    ttk.Label(win, text="Fe₂O₃").grid(column=0, row=6)
    ttk.Label(win, text="CaO").grid(column=0, row=7)
    ttk.Label(win, text="MgO").grid(column=0, row=8)
    ttk.Label(win, text="SO₃").grid(column=0, row=9)
    ttk.Label(win, text="Na₂O").grid(column=0, row=10)
    ttk.Label(win, text="K₂O").grid(column=0, row=11)
    ttk.Label(win, text="P₂O₅").grid(column=0, row=12)
    ttk.Label(win, text="Mn₂O₃").grid(column=0, row=13)
    ttk.Label(win, text="Cl").grid(column=0, row=14)
    ttk.Label(win, text="Free CaO").grid(column=0, row=15)
    ttk.Label(win, text="IR").grid(column=0, row=16)
    ttk.Label(win, text="Liter weight").grid(column=0, row=17)
    drop = OptionMenu(win, MaterialName, *materialOptions).grid(column=1, row=1)
    ttk.Entry(win, width=12, textvariable=Moisture).grid(column=1, row=2)
    ttk.Entry(win, width=12, textvariable=LOI).grid(column=1, row=3)
    ttk.Entry(win, width=12, textvariable=SiO2).grid(column=1, row=4)
    ttk.Entry(win, width=12, textvariable=Al2O3).grid(column=1, row=5)
    ttk.Entry(win, width=12, textvariable=Fe2O3).grid(column=1, row=6)
    ttk.Entry(win, width=12, textvariable=CaO).grid(column=1, row=7)
    ttk.Entry(win, width=12, textvariable=MgO).grid(column=1, row=8)
    ttk.Entry(win, width=12, textvariable=SO3).grid(column=1, row=9)
    ttk.Entry(win, width=12, textvariable=Na2O).grid(column=1, row=10)
    ttk.Entry(win, width=12, textvariable=K2O).grid(column=1, row=11)
    ttk.Entry(win, width=12, textvariable=P2O5).grid(column=1, row=12)
    ttk.Entry(win, width=12, textvariable=Mn2O3).grid(column=1, row=13)
    ttk.Entry(win, width=12, textvariable=Cl).grid(column=1, row=14)
    ttk.Entry(win, width=12, textvariable=FreeCaO).grid(column=1, row=15)
    ttk.Entry(win, width=12, textvariable=IR).grid(column=1, row=16)
    ttk.Entry(win, width=12, textvariable=Literweight).grid(column=1, row=17)

    ttk.Label(win, text="Date of Analysis").grid(column=2, row=0)
    ttk.Label(win, text="Source Name").grid(column=2, row=1)
    ttk.Label(win, text="Blaine").grid(column=2, row=2)
    ttk.Label(win, text="Residue +45 μm").grid(column=2, row=3)
    ttk.Label(win, text="Residue +90 μm").grid(column=2, row=4)
    ttk.Label(win, text="Residue +212 μm").grid(column=2, row=5)
    ttk.Label(win, text="Normal Consistency").grid(column=2, row=6)
    ttk.Label(win, text="LC Expansion").grid(column=2, row=7)
    ttk.Label(win, text="AC Expansion").grid(column=2, row=8)
    ttk.Label(win, text="Initial Setting Time").grid(column=2, row=9)
    ttk.Label(win, text="Final Setting Time").grid(column=2, row=10)
    ttk.Label(win, text="1 Day Strength").grid(column=2, row=11)
    ttk.Label(win, text="3 Days Strength").grid(column=2, row=12)
    ttk.Label(win, text="7 Days Strength").grid(column=2, row=13)
    ttk.Label(win, text="28 Days Strength").grid(column=2, row=14)
    ttk.Label(win, text="LR Strength").grid(column=2, row=15)
    ttk.Label(win, text="Drying Shrinkage").grid(column=2, row=16)
    ttk.Label(win, text="Replacement Test").grid(column=2, row=17)


    DoA = Calendar(win, selectmode='day', date_pattern="d-m-y")
    DoA.grid(column=3, row=0)
    ttk.Entry(win, width=12, textvariable=SourceName).grid(column=3, row=1)
    ttk.Entry(win, width=12, textvariable=Blaine).grid(column=3, row=2)
    ttk.Entry(win, width=12, textvariable=Residue1m).grid(column=3, row=3)
    ttk.Entry(win, width=12, textvariable=Residue2m).grid(column=3, row=4)
    ttk.Entry(win, width=12, textvariable=Residue3m).grid(column=3, row=5)
    ttk.Entry(win, width=12, textvariable=NormalConsistency).grid(column=3, row=6)
    ttk.Entry(win, width=12, textvariable=LCExpansion).grid(column=3, row=7)
    ttk.Entry(win, width=12, textvariable=ACExpansion).grid(column=3, row=8)
    ttk.Entry(win, width=12, textvariable=InitialSettingTime).grid(column=3, row=9)
    ttk.Entry(win, width=12, textvariable=FinalSettingTime).grid(column=3, row=10)
    ttk.Entry(win, width=12, textvariable=DayStrength1).grid(column=3, row=11)
    ttk.Entry(win, width=12, textvariable=DaysStrength3).grid(column=3, row=12)
    ttk.Entry(win, width=12, textvariable=DaysStrength7).grid(column=3, row=13)
    ttk.Entry(win, width=12, textvariable=DaysStrength28).grid(column=3, row=14)
    ttk.Entry(win, width=12, textvariable=LRStrength).grid(column=3, row=15)
    ttk.Entry(win, width=12, textvariable=DryingShrinkage).grid(column=3, row=16)
    ttk.Entry(win, width=12, textvariable=ReplacementTest).grid(column=3, row=17)

    timeDrop = OptionMenu(win, Time, *timeOptions).grid(column=4, row=0)
    ttk.Label(win, text="Party Name").grid(column=4, row=1)
    ttk.Label(win, text="+ 50.0 mm").grid(column=4, row=2)
    ttk.Label(win, text="+ 25.0 mm").grid(column=4, row=3)
    ttk.Label(win, text="+ 12.5 mm").grid(column=4, row=4)
    ttk.Label(win, text="+ 6.5 mm").grid(column=4, row=5)
    ttk.Label(win, text="+ 4.0 mm").grid(column=4, row=6)
    ttk.Label(win, text="+ 2.0 mm").grid(column=4, row=7)
    ttk.Label(win, text="+ 1.0 mm").grid(column=4, row=8)
    ttk.Label(win, text="+ 0.5 mm").grid(column=4, row=9)
    ttk.Label(win, text="Surface Moisture").grid(column=4, row=10)
    ttk.Label(win, text="Inherent Moisture").grid(column=4, row=11)
    ttk.Label(win, text="Voltaile Matter").grid(column=4, row=12)
    ttk.Label(win, text="Fixed Carbon").grid(column=4, row=13)
    ttk.Label(win, text="Ash").grid(column=4, row=14)
    ttk.Label(win, text="NCV").grid(column=4, row=15)
    ttk.Label(win, text="GCV").grid(column=4, row=16)
    ttk.Label(win, text="pH").grid(column=4, row=17)
    ttk.Entry(win, width=12, textvariable=PartyName).grid(column=5, row=1)
    ttk.Entry(win, width=12, textvariable=mm50).grid(column=5, row=2)
    ttk.Entry(win, width=12, textvariable=mm25).grid(column=5, row=3)
    ttk.Entry(win, width=12, textvariable=mm12).grid(column=5, row=4)
    ttk.Entry(win, width=12, textvariable=mm6).grid(column=5, row=5)
    ttk.Entry(win, width=12, textvariable=mm4).grid(column=5, row=6)
    ttk.Entry(win, width=12, textvariable=mm2).grid(column=5, row=7)
    ttk.Entry(win, width=12, textvariable=mm1).grid(column=5, row=8)
    ttk.Entry(win, width=12, textvariable=mm0).grid(column=5, row=9)
    ttk.Entry(win, width=12, textvariable=SurfaceMoisture).grid(column=5, row=10)
    ttk.Entry(win, width=12, textvariable=InherentMoisture).grid(column=5, row=11)
    ttk.Entry(win, width=12, textvariable=VoltaileMatter).grid(column=5, row=12)
    ttk.Entry(win, width=12, textvariable=FixedCarbon).grid(column=5, row=13)
    ttk.Entry(win, width=12, textvariable=Ash).grid(column=5, row=14)
    ttk.Entry(win, width=12, textvariable=NCV).grid(column=5, row=15)
    ttk.Entry(win, width=12, textvariable=GCV).grid(column=5, row=16)
    ttk.Entry(win, width=12, textvariable=pH).grid(column=5, row=17)

    ttk.Label(win, text="Remarks").grid(column=0, row=18)
    ttk.Entry(win, width=28, textvariable=remarks).grid(column=1, row=18)


MaterialName = tk.StringVar()
MaterialName.set("Select")
Moisture = tk.StringVar()
LOI = tk.StringVar()
SiO2 = tk.StringVar()
Al2O3 = tk.StringVar()
Fe2O3 = tk.StringVar()
CaO = tk.StringVar()
MgO = tk.StringVar()
SO3 = tk.StringVar()
Na2O = tk.StringVar()
K2O = tk.StringVar()
P2O5 = tk.StringVar()
Mn2O3 = tk.StringVar()
Cl = tk.StringVar()
FreeCaO = tk.StringVar()
IR = tk.StringVar()
Literweight = tk.StringVar()

SourceName = tk.StringVar()
Blaine = tk.StringVar()
Residue1m = tk.StringVar()
Residue2m = tk.StringVar()
Residue3m = tk.StringVar()
NormalConsistency = tk.StringVar()
LCExpansion = tk.StringVar()
ACExpansion = tk.StringVar()
InitialSettingTime = tk.StringVar()
FinalSettingTime = tk.StringVar()
DayStrength1 = tk.StringVar()
DaysStrength3 = tk.StringVar()
DaysStrength7 = tk.StringVar()
DaysStrength28 = tk.StringVar()
LRStrength = tk.StringVar()
DryingShrinkage = tk.StringVar()
ReplacementTest = tk.StringVar()

Time = tk.StringVar()
Time.set("Select")
PartyName = tk.StringVar()
mm50 = tk.StringVar()
mm25 = tk.StringVar()
mm12 = tk.StringVar()
mm6 = tk.StringVar()
mm4 = tk.StringVar()
mm2 = tk.StringVar()
mm1 = tk.StringVar()
mm0 = tk.StringVar()
SurfaceMoisture = tk.StringVar()
InherentMoisture = tk.StringVar()
VoltaileMatter = tk.StringVar()
FixedCarbon = tk.StringVar()
Ash = tk.StringVar()
NCV = tk.StringVar()
GCV = tk.StringVar()
pH = tk.StringVar()

remarks = tk.StringVar()

timeSelection()


def filterContents():
    print(filterDateFrom.get_date(), filterDateTo.get_date())
    c.execute("SELECT * FROM clinkerData WHERE Date_of_Sample < {} and Date_of_Sample >= {}".format(filterDateFrom, filterDateTo))

    for item in c.fetchall():
        print(item[0] + "\t " + item[1])



def createNewWindow():
    win = tk.Tk()  # Application Name
    win.title("Filter")  # Label

    ttk.Label(win, text="From").grid(column=0, row=0)
    ttk.Label(win, text="To").grid(column=2, row=0)

    global filterDateFrom, filterDateTo

    filterDateFrom = Calendar(win, selectmode='day', date_pattern="d-m-y")
    filterDateFrom.grid(column=1, row=0)
    filterDateTo = Calendar(win, selectmode='day', date_pattern="d-m-y")
    filterDateTo.grid(column=3, row=0)

    filterBtn = ttk.Button(win, text="submit", command=filterContents).grid(column=2, row=18)


button = ttk.Button(win, text="submit", command=listofcommands).grid(column=2, row=18)
loadbtn = ttk.Button(win, text="Load data", command=createNewWindow).grid(column=3, row=18)

win.mainloop()


# Data types:
# null
# integer
# real
# text
# blob

# query the database
# c.execute("""SELECT * FROM clinkerData""")
# print(c.fetchone()[0])
# c.fetchmany(3)
# print(c.fetchall())

for item in c.fetchall():
    print(item[0] + "\t " + item[1] + "\t " + str(item[2]))


def savingExcelData():
    pass


savingExcelData()



# commit our database connection
clin.commit()

# close out connection
clin.close()
